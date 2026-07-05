#!/usr/bin/env python3
"""Baixa os dados brutos de fiscalização do abastecimento da ANP (xlsx),
valida, converte para CSV no formato consumido pelo app octano e grava
data/base_anp.csv + data/metadata.json.

Sai com código != 0 em qualquer problema (download, header inesperado,
contagem de linhas suspeita), o que faz o workflow falhar e disparar a
notificação de erro.
"""

import csv
import hashlib
import io
import json
import sys
import time
import urllib.request
import zipfile
from datetime import datetime, timezone
from pathlib import Path

import xlrd
from openpyxl import load_workbook

SOURCE_URL = (
    "https://www.gov.br/anp/pt-br/centrais-de-conteudo/paineis-dinamicos-da-anp/"
    "arquivos-dados-brutos-do-painel-dinamico-da-fiscalizacao-do-abastecimento-da-sfi/"
    "dados-fisc-a-partir-2019.xlsx"
)

# API pública de revendedores da ANP: cadastro oficial com CNPJ, razão social,
# bandeira e COORDENADAS de cada posto — usada pelo app para validar
# geograficamente o cruzamento Google ⇄ ANP.
CADASTRO_API = "https://revendedoresapi.anp.gov.br/v1/combustivel?numeropagina={page}"
CADASTRO_MIN_ROWS = 30_000  # a base tem ~46 mil postos; menos que isso é problema

EXPECTED_HEADER = [
    "UF",
    "Município",
    "Bairro",
    "ENDEREÇO",
    "CNPJ/CPF",
    "Agente Econômico",
    "Segmento Fiscalizado",
    "DATA DO DF",
    "Número do Documento",
    "Procedimento de Fiscalização",
    "Resultado",
]

# A base de abr/2025 tinha ~225 mil linhas e só cresce (dados de 2019 em diante).
# Se a ANP publicar um arquivo muito menor, algo mudou no formato: melhor falhar.
MIN_ROWS = 200_000

# --- Procon-SP: Portal da Transparência "Empresas Autuadas" (multas com ---
# --- processo administrativo encerrado e não pagas, por CNPJ).          ---
# Dados públicos de transparência ativa (LAI); coleta 1x/semana, volume
# baixo, com identificação no User-Agent. Só pessoas jurídicas.
PROCON_SP_API = (
    "https://sistemas.procon.sp.gov.br/transparencia/empresas_autuadas/"
    "divida_grupo_empresa.php"
)
# O app só precisa de postos de combustível: filtramos por CNPJ presente no
# cadastro de revendedores da ANP OU por palavra-chave no nome (o cadastro
# não cobre postos já fechados, que ainda interessam ao histórico).
PROCON_FUEL_KEYWORDS = ("POSTO", "COMBUST", "ABASTECEDORA", "CENTRO AUTOMOTIVO")
PROCON_MIN_TOTAL = 3_000    # a lista completa tem ~5 mil empresas
PROCON_MIN_FUEL = 300       # ~850 são postos; menos que isso é problema


# --- Preços de DISTRIBUIÇÃO (semanal, desde ago/2020): Brasil e por UF. ---
PRECOS_DIST_URLS = [
    ("BRASIL", "https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/"
               "precos/pdc/semanal/combustiveis-liquidos-brasil.xlsx"),
    ("ESTADOS", "https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/"
                "precos/pdc/semanal/combustiveis-liquidos-estados.xlsx"),
]
PRECOS_DIST_MIN_ROWS = 20_000  # jul/2026: ~42 mil (Brasil + 27 UFs × 5 produtos)

# --- Preços de PRODUTORES/importadores (semanal, desde 2013): por região. ---
PRECOS_PRODUTORES_URL = (
    "https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/"
    "precos/ppidp/precos-medios-ponderados-semanais-2013.xls"
)
# A planilha tem 21 produtos (asfalto, nafta, QAV...); publicamos só os
# relevantes para o consumidor do app.
PRODUTORES_PRODUTOS = (
    "Gasolina A Comum",
    "Gasolina A Premium",
    "Óleo Diesel S-10",
    "Óleo Diesel S-500",
    "Biodiesel B-100",
    "Gás Liquefeito de Petróleo",
)
PRECOS_PROD_MIN_ROWS = 3_000

# --- Reclamações Fundamentadas (PowerBI público do Procon-SP):           ---
# --- CNPJs de empresas com reclamações de consumidores procedentes.      ---
# Dashboard: https://www.procon.sp.gov.br/empresas-reclamadas/
# A chave pública do relatório é estável enquanto o Procon não republicar.
# A API limita cada resposta a 30 mil linhas; a lista completa (~394 mil
# CNPJs) é percorrida com os RestartTokens que a própria resposta devolve.
RECLAMADAS_POWERBI_URL = (
    "https://wabi-brazil-south-b-primary-api.analysis.windows.net/"
    "public/reports/querydata?synchronous=true"
)
RECLAMADAS_RESOURCE_KEY = "7f552da0-bf9d-4c97-be36-fbea30e0dd3d"
# Colunas de nome da tabela "account" usadas no filtro de postos fechados.
RECLAMADAS_NAME_COLUMNS = ("Razão Social", "Nome Fantasia(Pesquisa)")
RECLAMADAS_MIN_TOTAL = 100_000  # a lista completa tem ~394 mil CNPJs
RECLAMADAS_MIN_FUEL = 300       # ~8 mil batem no filtro de postos

# --- IPEM-SP "Bombas Seguras" (PowerBI público): postos com bombas         ---
# --- certificadas antifraude (selo "Bombas Certificadas IPEM" do app).     ---
# Painel: https://www.ipem.sp.gov.br/bombasegura
# O modelo BaseSgi expõe o CNPJ do posto; filtramos por isvalid='Autorizado'.
# Base pequena (SP inteiro tem ~200 postos), uma resposta única já cobre.
IPEM_POWERBI_URL = RECLAMADAS_POWERBI_URL
IPEM_RESOURCE_KEY = "82bf5d55-292e-4cf4-8fc6-61605c816c8b"
IPEM_DATASET_ID = "8971c0f9-307b-46b8-a211-5eba41585040"
IPEM_REPORT_ID = "eda4f821-4a5f-441c-ab5f-9a978f53a6a4"
IPEM_VISUAL_ID = "2a7ae5e11f58d1550f50"
IPEM_MODEL_ID = 5200277
IPEM_MIN = 100  # jul/2026: 207 postos autorizados; menos que isso é problema

# --- SINDEC: Cadastro Nacional de Reclamações Fundamentadas (Senacon).      ---
# --- Reclamações fundamentadas de TODOS os Procons do Brasil (o do DF       ---
# --- incluso), por CNPJ, com CNAE e status de atendimento. Filtramos postos ---
# --- de combustível (CNAE 4731* ou palavra-chave) e agregamos por CNPJ com  ---
# --- índice de resolutividade (atendidas / total).                          ---
# Fonte: https://dados.mj.gov.br/dataset/cadastro-nacional-de-reclamacoes-fundamentadas-procons-sindec
# Um recurso (URL) por ano; o formato varia (CSV direto, CSV/XLSX em zip).
_SINDEC_BASE = ("https://dados.mj.gov.br/dataset/"
                "8ff7032a-d6db-452b-89f1-d860eb6965ff/resource")
SINDEC_YEARS = {
    2015: f"{_SINDEC_BASE}/f9d3a86b-2f58-435e-a100-77926af0469a/download/reclamacoes-fundamentadas-sindec-2015.zip",
    2016: f"{_SINDEC_BASE}/4d055554-0595-47ce-b3d4-97c11f33e143/download/reclamacoes-fundamentadas-sindec-2016v2.zip",
    2017: f"{_SINDEC_BASE}/8d400ac5-6aad-4ad9-a33a-74ca40f2242e/download/cnrf2017.zip",
    2018: f"{_SINDEC_BASE}/a37f1be2-bdb0-4191-9563-015a40e65434/download/reclamacoes-fundamentadas-2018-em-zip.zip",
    2019: f"{_SINDEC_BASE}/c2cce323-24c2-4430-8918-e24b2966213c/download/crf2019-dados-abertos.zip",
    2020: f"{_SINDEC_BASE}/094e22a2-82f0-448f-b377-1985cbe99ec8/download/cnrf2020.zip",
    2021: f"{_SINDEC_BASE}/a709fa60-00d2-47db-a86a-97eb10fa62ff/download/cnrf2021.csv",
    2022: f"{_SINDEC_BASE}/bc23b54c-18e7-4ed8-b7f5-205434ac5719/download/crf2022dados-abertos.csv",
    2023: f"{_SINDEC_BASE}/e0c5eaea-ace1-457d-a945-9645644d2783/download/cnrf2023dadosabertos.zip",
    2024: f"{_SINDEC_BASE}/34a47b5c-5577-421a-bf9b-2dbbfc371d65/download/cnrfdadosabertos2024.zip",
}
SINDEC_FUEL_KEYWORDS = ("POSTO", "COMBUST", "DERIVADOS DE PETROLEO", "ABASTEC")
SINDEC_MIN = 500  # nacional, 2015+: alguns milhares de postos; menos é problema

DATA_DIR = Path(__file__).resolve().parent.parent / "data"


def mask_cpf(value: str) -> str:
    """Mascara CPFs de revendedores pessoa física (LGPD): a coluna da ANP é
    "CNPJ/CPF" e ~1.200 registros trazem CPF completo. CNPJ (14 dígitos, dado
    de empresa) passa intacto; CPF (11 dígitos) mantém só os 6 dígitos do
    meio — suficiente para conferência, sem redistribuir o documento."""
    digits = "".join(ch for ch in value if ch.isdigit())
    if len(digits) == 11:
        return f"***{digits[3:9]}**"
    return value


def format_cell(value, is_date_column: bool) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        # Mesmo formato do CSV histórico do app: M/D/YYYY sem zero à esquerda.
        return f"{value.month}/{value.day}/{value.year}"
    text = str(value)
    if is_date_column and text and "/" not in text:
        # Data que veio como texto/serial inesperado: falha em vez de corromper.
        raise ValueError(f"Valor de data não reconhecido: {text!r}")
    # O parser de CSV do app é linha-a-linha e não trata aspas escapadas:
    # remove quebras de linha e aspas duplas de dentro das células.
    return text.replace("\r", " ").replace("\n", " ").replace('"', "'").strip()


def fetch_cadastro() -> bytes:
    """Baixa o cadastro completo de postos da API de revendedores (paginada,
    5.000 registros por página) e devolve um CSV `;`-separado com CNPJ,
    razão social, bandeira, UF, município, latitude e longitude."""
    lines = ["CNPJ;RAZAOSOCIAL;BANDEIRA;UF;MUNICIPIO;LATITUDE;LONGITUDE"]
    count = 0
    page = 1
    while page <= 30:  # trava de segurança: hoje são ~10 páginas
        url = CADASTRO_API.format(page=page)
        print(f"Cadastro: página {page}...")
        request = urllib.request.Request(url, headers={
            "User-Agent": "octano-anp-data/1.0", "Accept": "application/json",
        })
        with urllib.request.urlopen(request, timeout=300) as response:
            payload = json.loads(response.read())
        rows = payload.get("data") or []
        if not rows:
            break
        for row in rows:
            def clean(value):
                return str(value or "").replace(";", ",").replace("\n", " ").strip()
            lines.append(";".join([
                clean(row.get("cnpj")),
                clean(row.get("razaoSocial")),
                clean(row.get("distribuidora")),
                clean(row.get("uf")),
                clean(row.get("municipio")),
                clean(row.get("latitude")),
                clean(row.get("longitude")),
            ]))
            count += 1
        page += 1

    print(f"Cadastro: {count} postos")
    if count < CADASTRO_MIN_ROWS:
        raise RuntimeError(
            f"Cadastro com apenas {count} postos (mínimo {CADASTRO_MIN_ROWS}) — abortando"
        )
    return ("\n".join(lines) + "\n").encode("utf-8")


def download(url: str, attempts: int = 4) -> bytes:
    """Download com retry exponencial: o gov.br falha de forma transitória
    com alguma frequência ("Network is unreachable") e o workflow é semanal —
    melhor insistir alguns minutos do que falhar a atualização inteira."""
    last_error: Exception | None = None
    for attempt in range(1, attempts + 1):
        try:
            request = urllib.request.Request(
                url, headers={"User-Agent": "octano-anp-data/1.0"}
            )
            with urllib.request.urlopen(request, timeout=300) as response:
                return response.read()
        except Exception as error:  # noqa: BLE001 - urllib levanta tipos variados
            last_error = error
            if attempt < attempts:
                wait = 30 * attempt
                print(f"Tentativa {attempt} falhou ({error}); nova em {wait}s...")
                time.sleep(wait)
    raise RuntimeError(f"Download falhou após {attempts} tentativas: {last_error}")


def _procon_get(params: str) -> dict:
    payload = download(f"{PROCON_SP_API}?{params}", attempts=3)
    data = json.loads(payload)
    if data.get("Result") != "OK":
        raise RuntimeError(f"Procon-SP respondeu {data.get('Result')!r} para {params}")
    return data


def _is_fuel_name(name: str) -> bool:
    upper = (name or "").upper()
    return any(keyword in upper for keyword in PROCON_FUEL_KEYWORDS)


def fetch_procon_sp(cadastro_cnpjs: set) -> bytes:
    """Baixa as empresas autuadas pelo Procon-SP e devolve um CSV apenas com
    postos de combustível, uma linha por processo:
    CNPJ;RAZAO_SOCIAL;PROCESSO;STATUS;VALOR_MULTA

    Importante: o portal lista multas de processos ENCERRADOS e não pagas
    (boa parte inscrita na Dívida Ativa) — não são meras acusações, mas o
    app ainda assim exibe como registro público atribuído à fonte."""
    data = _procon_get("action=lista_grupo_empresa&busca=&jtStartIndex=0&jtPageSize=20000")
    records = data["Records"]
    print(f"Procon-SP: {len(records)} empresas/grupos na lista")
    if len(records) < PROCON_MIN_TOTAL:
        raise RuntimeError(f"Procon-SP com só {len(records)} registros — abortando")

    # Empresa individual: CNPJ direto. Grupo com nome de posto: expande os
    # CNPJs dos membros (raros — hoje são ~8 grupos).
    selected = {}  # cnpj -> nome
    for record in records:
        name = record.get("nome_grupo_empresa") or ""
        cnpj = record.get("cnpj")
        if record.get("eh_grupo") == "1":
            if _is_fuel_name(name):
                group = _procon_get(
                    f"action=lista_empresas_do_grupo&id={record['id_grupo_empresa']}"
                )
                for member in group["Records"]:
                    member_cnpj = member.get("cnpj")
                    if member_cnpj:
                        selected[member_cnpj] = member.get("fornecedor_nome") or name
                time.sleep(0.1)
        elif cnpj and (cnpj in cadastro_cnpjs or _is_fuel_name(name)):
            selected[cnpj] = name

    print(f"Procon-SP: {len(selected)} postos de combustível identificados")
    if len(selected) < PROCON_MIN_FUEL:
        raise RuntimeError(f"Só {len(selected)} postos no Procon-SP — abortando")

    lines = ["CNPJ;RAZAO_SOCIAL;PROCESSO;STATUS;VALOR_MULTA"]
    count = 0
    for index, (cnpj, name) in enumerate(sorted(selected.items())):
        processos = _procon_get(f"action=lista_processos_da_empresa&cnpj={cnpj}")
        for proc in processos["Records"]:
            def clean(value):
                return str(value or "").replace(";", ",").replace("\n", " ").strip()
            lines.append(";".join([
                cnpj,
                clean(name),
                clean(proc.get("processo")),
                clean(proc.get("status_nome")),
                clean(proc.get("valor_multa")),
            ]))
            count += 1
        if index % 100 == 0:
            print(f"Procon-SP: processos de {index}/{len(selected)} postos...")
        time.sleep(0.1)

    print(f"Procon-SP: {count} processos de {len(selected)} postos")
    return ("\n".join(lines) + "\n").encode("utf-8")


def _normalize_cnpj(raw: str) -> str:
    """Normaliza um CNPJ vindo do PowerBI: remove formatação, preenche com
    zeros à esquerda até 14 dígitos. Retorna string vazia se inválido."""
    digits = "".join(ch for ch in raw if ch.isdigit())
    if not digits or len(digits) > 14:
        return ""
    return digits.zfill(14)


def _reclamadas_post(body: dict) -> dict:
    """POST na API pública do PowerBI com retry (mesma política dos downloads
    do gov.br: falhas transitórias não devem derrubar a coleta semanal)."""
    payload = json.dumps(body).encode("utf-8")
    headers = {
        "Content-Type": "application/json;charset=UTF-8",
        "Accept": "application/json, text/plain, */*",
        "X-PowerBI-ResourceKey": RECLAMADAS_RESOURCE_KEY,
        "User-Agent": "octano-anp-data/1.0",
        "Origin": "https://app.powerbi.com",
        "Referer": "https://app.powerbi.com/",
    }
    last_error: Exception | None = None
    for attempt in range(1, 4):
        try:
            request = urllib.request.Request(
                RECLAMADAS_POWERBI_URL, data=payload, headers=headers, method="POST"
            )
            with urllib.request.urlopen(request, timeout=120) as response:
                return json.loads(response.read())
        except Exception as error:  # noqa: BLE001
            last_error = error
            if attempt < 3:
                wait = 15 * attempt
                print(f"Tentativa {attempt} falhou ({error}); nova em {wait}s...")
                time.sleep(wait)
    raise RuntimeError(f"PowerBI Reclamadas falhou após 3 tentativas: {last_error}")


def _reclamadas_fetch_cnpjs(extra_where: list | None = None) -> set:
    """Consulta a coluna account.CNPJ do relatório (agrupada, ou seja, valores
    distintos) e devolve o conjunto de CNPJs normalizados. A API entrega no
    máximo 30 mil linhas por resposta; quando há mais, devolve RestartTokens
    ("RT") que retomam a listagem de onde parou — seguimos até o fim."""
    where = [{"Condition": {"Not": {"Expression": {"In": {
        "Expressions": [{"Column": {
            "Expression": {"SourceRef": {"Source": "a"}},
            "Property": "CNPJ",
        }}],
        "Values": [[{"Literal": {"Value": "null"}}]],
    }}}}}]
    where += extra_where or []

    cnpjs: set = set()
    restart_tokens = None
    for _page in range(40):  # trava de segurança: hoje são 14 páginas
        window: dict = {"Count": 30000}
        if restart_tokens:
            window["RestartTokens"] = restart_tokens
        body = {
            "version": "1.0.0",
            "queries": [{
                "Query": {"Commands": [{"SemanticQueryDataShapeCommand": {
                    "Query": {
                        "Version": 2,
                        "From": [{"Name": "a", "Entity": "account", "Type": 0}],
                        "Select": [{
                            "Column": {
                                "Expression": {"SourceRef": {"Source": "a"}},
                                "Property": "CNPJ",
                            },
                            "Name": "account.CNPJ",
                        }],
                        "Where": where,
                    },
                    "Binding": {
                        "Primary": {"Groupings": [{"Projections": [0]}]},
                        "DataReduction": {"DataVolume": 3, "Primary": {"Window": window}},
                        "Version": 1,
                    },
                    "ExecutionMetricsKind": 1,
                }}]},
                "QueryId": "",
                "ApplicationContext": {
                    "DatasetId": "5410ab72-d62e-4976-bf13-c8039a70b5ee",
                    "Sources": [{
                        "ReportId": "98fdc96d-085b-4ca1-ac68-d29bfed78e8f",
                        "VisualId": "3e04fd14718d00c9ce9d",
                    }],
                },
            }],
            "cancelQueries": [],
            "modelId": 3819509,
        }

        data = _reclamadas_post(body)
        try:
            dataset = data["results"][0]["result"]["data"]["dsr"]["DS"][0]
            for group in dataset.get("PH", []):
                for row in group.get("DM0", []):
                    normalized = _normalize_cnpj(str(row.get("G0", "")))
                    if len(normalized) == 14:
                        cnpjs.add(normalized)
            restart_tokens = dataset.get("RT")
        except (KeyError, IndexError, TypeError) as err:
            raise RuntimeError(
                f"Estrutura do PowerBI mudou — não foi possível extrair CNPJs: {err}"
            ) from err

        if not restart_tokens:
            return cnpjs
        time.sleep(0.5)
    raise RuntimeError("PowerBI Reclamadas não terminou em 40 páginas — abortando")


def fetch_reclamadas_procon_sp(cadastro_cnpjs: set) -> bytes:
    """Extrai CNPJs de empresas reclamadas do PowerBI público do Procon-SP
    (Reclamações Fundamentadas) e filtra para postos de combustível com a
    mesma política do fetch_procon_sp: CNPJ presente no cadastro de
    revendedores da ANP OU palavra-chave de posto no nome (razão social /
    nome fantasia), que cobre postos já fechados fora do cadastro.

    Devolve um CSV simples: CNPJ (uma coluna, um CNPJ normalizado por linha).
    Fonte: https://www.procon.sp.gov.br/empresas-reclamadas/"""
    print("Reclamadas Procon-SP (PowerBI): baixando lista completa de CNPJs...")
    todos = _reclamadas_fetch_cnpjs()
    print(f"Reclamadas Procon-SP: {len(todos)} CNPJs na lista completa")
    if len(todos) < RECLAMADAS_MIN_TOTAL:
        raise RuntimeError(
            f"Lista completa com só {len(todos)} CNPJs (mínimo {RECLAMADAS_MIN_TOTAL}) — abortando"
        )

    normalizado_cadastro = {_normalize_cnpj(cnpj) for cnpj in cadastro_cnpjs}
    fuel = todos & normalizado_cadastro
    print(f"Reclamadas Procon-SP: {len(fuel)} CNPJs no cadastro da ANP")

    # Postos fechados não constam mais no cadastro: busca no servidor por
    # palavra-chave nas colunas de nome (Contains é case-insensitive).
    for column in RECLAMADAS_NAME_COLUMNS:
        for keyword in PROCON_FUEL_KEYWORDS:
            extra = [{"Condition": {"Contains": {
                "Left": {"Column": {
                    "Expression": {"SourceRef": {"Source": "a"}},
                    "Property": column,
                }},
                "Right": {"Literal": {"Value": f"'{keyword}'"}},
            }}}]
            matched = _reclamadas_fetch_cnpjs(extra)
            fuel |= matched
            print(f"  - {column} contém {keyword!r}: {len(matched)} CNPJs")
            time.sleep(0.5)

    print(f"Reclamadas Procon-SP: {len(fuel)} postos de combustível")
    if len(fuel) < RECLAMADAS_MIN_FUEL:
        raise RuntimeError(
            f"Só {len(fuel)} postos nas Reclamadas (mínimo {RECLAMADAS_MIN_FUEL}) — abortando"
        )

    lines = ["CNPJ"] + sorted(fuel)
    return ("\n".join(lines) + "\n").encode("utf-8")


def _ipem_post(body: dict) -> dict:
    """POST na API pública do PowerBI do IPEM-SP com retry (mesma política do
    _reclamadas_post, com a resource key do painel Bombas Seguras)."""
    payload = json.dumps(body).encode("utf-8")
    headers = {
        "Content-Type": "application/json;charset=UTF-8",
        "Accept": "application/json, text/plain, */*",
        "X-PowerBI-ResourceKey": IPEM_RESOURCE_KEY,
        "User-Agent": "octano-anp-data/1.0",
        "Origin": "https://app.powerbi.com",
        "Referer": "https://app.powerbi.com/",
    }
    last_error: Exception | None = None
    for attempt in range(1, 4):
        try:
            request = urllib.request.Request(
                IPEM_POWERBI_URL, data=payload, headers=headers, method="POST"
            )
            with urllib.request.urlopen(request, timeout=120) as response:
                return json.loads(response.read())
        except Exception as error:  # noqa: BLE001
            last_error = error
            if attempt < 3:
                wait = 15 * attempt
                print(f"Tentativa {attempt} falhou ({error}); nova em {wait}s...")
                time.sleep(wait)
    raise RuntimeError(f"PowerBI IPEM falhou após 3 tentativas: {last_error}")


def fetch_ipem_bombas_seguras() -> bytes:
    """Extrai os CNPJs dos postos com bombas certificadas (autorizadas) do
    painel público "Bombas Seguras" do IPEM-SP e devolve um JSON: array de
    CNPJs normalizados (14 dígitos), no formato que o IPEMProvider do app lê.
    Fonte: https://www.ipem.sp.gov.br/bombasegura"""
    print("IPEM Bombas Seguras (PowerBI): baixando postos autorizados...")
    body = {
        "version": "1.0.0",
        "queries": [{
            "Query": {"Commands": [{"SemanticQueryDataShapeCommand": {
                "Query": {
                    "Version": 2,
                    "From": [{"Name": "b", "Entity": "BaseSgi", "Type": 0}],
                    "Select": [{
                        "Column": {
                            "Expression": {"SourceRef": {"Source": "b"}},
                            "Property": "CNPJ",
                        },
                        "Name": "BaseSgi.CNPJ",
                    }],
                    "Where": [{"Condition": {"In": {
                        "Expressions": [{"Column": {
                            "Expression": {"SourceRef": {"Source": "b"}},
                            "Property": "isvalid",
                        }}],
                        "Values": [[{"Literal": {"Value": "'Autorizado'"}}]],
                    }}}],
                },
                "Binding": {
                    "Primary": {"Groupings": [{"Projections": [0]}]},
                    "DataReduction": {"DataVolume": 3, "Primary": {"Window": {"Count": 5000}}},
                    "Version": 1,
                },
                "ExecutionMetricsKind": 1,
            }}]},
            "QueryId": "",
            "ApplicationContext": {
                "DatasetId": IPEM_DATASET_ID,
                "Sources": [{"ReportId": IPEM_REPORT_ID, "VisualId": IPEM_VISUAL_ID}],
            },
        }],
        "cancelQueries": [],
        "modelId": IPEM_MODEL_ID,
    }

    data = _ipem_post(body)
    try:
        dataset = data["results"][0]["result"]["data"]["dsr"]["DS"][0]
    except (KeyError, IndexError, TypeError) as err:
        raise RuntimeError(f"Estrutura do PowerBI IPEM mudou: {err}") from err

    # Cada linha traz o CNPJ na chave "G0" (mesmo formato do PowerBI de
    # reclamadas); são valores distintos, uma linha por posto autorizado.
    cnpjs: set = set()
    for group in dataset.get("PH", []):
        for row in group.get("DM0", []):
            normalized = _normalize_cnpj(str(row.get("G0", "")))
            if len(normalized) == 14:
                cnpjs.add(normalized)

    print(f"IPEM Bombas Seguras: {len(cnpjs)} postos autorizados")
    if len(cnpjs) < IPEM_MIN:
        raise RuntimeError(
            f"Só {len(cnpjs)} postos no IPEM (mínimo {IPEM_MIN}) — abortando"
        )
    return (json.dumps(sorted(cnpjs), ensure_ascii=False) + "\n").encode("utf-8")


def _sindec_iter_rows(data: bytes, name: str):
    """Gera dicts {coluna: valor} de um arquivo do SINDEC, tratando as três
    formas em que a Senacon publica: CSV direto, CSV dentro de zip e XLSX
    dentro de zip. Encoding do CSV: utf-8 (com BOM) ou latin-1."""
    lower = name.lower()
    if lower.endswith(".zip"):
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            for member in archive.namelist():
                if member.endswith("/"):
                    continue
                yield from _sindec_iter_rows(archive.read(member), member)
        return
    if lower.endswith(".csv"):
        text = None
        for encoding in ("utf-8-sig", "latin-1"):
            try:
                text = data.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        reader = csv.DictReader(io.StringIO(text or ""), delimiter=";")
        reader.fieldnames = [(fn or "").lstrip("﻿") for fn in (reader.fieldnames or [])]
        yield from reader
        return
    if lower.endswith(".xlsx"):
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        header = [
            (str(cell).strip().lstrip("﻿") if cell is not None else "")
            for cell in next(rows)
        ]
        for row in rows:
            yield {
                col: (row[i] if i < len(row) else None)
                for i, col in enumerate(header)
            }


def fetch_reclamadas_sindec() -> bytes:
    """Baixa o Cadastro Nacional de Reclamações Fundamentadas (SINDEC/Senacon)
    de vários anos, filtra postos de combustível (CNAE 4731* ou palavra-chave
    no nome) e agrega por CNPJ, devolvendo um CSV nacional:
    CNPJ;RAZAO_SOCIAL;UFS;ATENDIDAS;NAO_ATENDIDAS;TOTAL;RESOLUTIVIDADE

    RESOLUTIVIDADE = atendidas / total (em %), o mesmo índice que os Procons
    publicam. Cobre todos os estados; o Procon-DF entra como parte do SINDEC.
    Fonte: dados.mj.gov.br (Ministério da Justiça)."""
    # cnpj -> [atendidas, nao_atendidas, nome, {UFs}]
    agg: dict = {}
    for year, url in sorted(SINDEC_YEARS.items()):
        print(f"SINDEC {year}: baixando...")
        try:
            payload = download(url, attempts=3)
        except Exception as error:  # noqa: BLE001 - um ano fora não derruba o resto
            print(f"AVISO: SINDEC {year} indisponível ({error}); pulando o ano")
            continue

        total = fuel = 0
        for row in _sindec_iter_rows(payload, url):
            total += 1
            cnae = "".join(ch for ch in str(row.get("CNAEPrincipal") or "") if ch.isdigit())
            razao = str(row.get("strRazaoSocial") or "")
            fantasia = str(row.get("strNomeFantasia") or "")
            is_fuel = cnae.startswith("4731") or any(
                keyword in (razao + fantasia).upper() for keyword in SINDEC_FUEL_KEYWORDS
            )
            if not is_fuel:
                continue

            cnpj = _normalize_cnpj(str(row.get("NumeroCNPJ") or ""))
            # Descarta CNPJ inválido/zerado (a base traz alguns "00000000000000").
            if len(cnpj) != 14 or int(cnpj) < 1_000_000:
                continue
            fuel += 1

            record = agg.setdefault(cnpj, [0, 0, "", set()])
            if str(row.get("Atendida") or "").strip().upper() == "S":
                record[0] += 1
            else:
                record[1] += 1
            record[2] = razao or fantasia or record[2]
            uf = str(row.get("UF") or "").strip().upper()
            if uf:
                record[3].add(uf)
        print(f"SINDEC {year}: {total} reclamações, {fuel} de postos")

    print(f"SINDEC: {len(agg)} postos distintos com reclamação fundamentada")
    if len(agg) < SINDEC_MIN:
        raise RuntimeError(
            f"Só {len(agg)} postos no SINDEC (mínimo {SINDEC_MIN}) — abortando"
        )

    def clean(value: str) -> str:
        return str(value or "").replace(";", ",").replace("\n", " ").strip()

    lines = ["CNPJ;RAZAO_SOCIAL;UFS;ATENDIDAS;NAO_ATENDIDAS;TOTAL;RESOLUTIVIDADE"]
    for cnpj, (atendidas, nao, nome, ufs) in sorted(agg.items()):
        total = atendidas + nao
        resolutividade = (100.0 * atendidas / total) if total else 0.0
        lines.append(";".join([
            cnpj,
            clean(nome),
            "/".join(sorted(ufs)),
            str(atendidas),
            str(nao),
            str(total),
            f"{resolutividade:.1f}",
        ]))
    return ("\n".join(lines) + "\n").encode("utf-8")


def fetch_precos_distribuicao() -> bytes:
    """Preços semanais de DISTRIBUIÇÃO (Brasil + por UF) num único CSV:
    INICIO;FIM;LOCAL;PRODUTO;UNIDADE;PRECO — LOCAL é "BRASIL" ou o nome da UF.
    Fonte: relatórios semanais de combustíveis líquidos da SDL/ANP."""
    lines = ["INICIO;FIM;LOCAL;PRODUTO;UNIDADE;PRECO"]
    count = 0

    for scope, url in PRECOS_DIST_URLS:
        print(f"Preços distribuição ({scope}): baixando...")
        workbook = load_workbook(io.BytesIO(download(url)), read_only=True, data_only=True)
        sheet = workbook.worksheets[0]

        col: dict[str, int] = {}
        price_index = -1
        for row in sheet.iter_rows(values_only=True):
            # As primeiras linhas são um cabeçalho institucional; a tabela
            # começa na linha cujo primeiro valor é "DATA INICIAL".
            if not col:
                if row and str(row[0]).strip().upper() == "DATA INICIAL":
                    names = [str(c).strip().upper() if c else "" for c in row]
                    col = {name: i for i, name in enumerate(names)}
                    # "PREÇO MÉDIO DISTRIBUIÇÃO" no arquivo de estados,
                    # "PREÇO MÉDIO DE DISTRIBUIÇÃO" no do Brasil.
                    price_index = next(
                        i for i, n in enumerate(names) if n.startswith("PREÇO MÉDIO")
                    )
                continue

            inicio = row[col["DATA INICIAL"]]
            fim = row[col["DATA FINAL"]]
            preco = row[price_index] if price_index < len(row) else None
            if not isinstance(inicio, datetime) or not isinstance(preco, (int, float)):
                continue

            local = str(row[col["ESTADO"]]).strip().upper() if "ESTADO" in col else "BRASIL"
            produto = str(row[col["PRODUTO"]]).strip().upper()
            unidade = (
                str(row[col["UNIDADE DE MEDIDA"]]).strip()
                if "UNIDADE DE MEDIDA" in col else ""
            )
            lines.append(";".join([
                inicio.strftime("%Y-%m-%d"),
                fim.strftime("%Y-%m-%d") if isinstance(fim, datetime) else "",
                local,
                produto,
                unidade,
                f"{float(preco):.3f}",
            ]))
            count += 1

        if not col:
            raise RuntimeError(f"Cabeçalho de preços não encontrado em {url}")

    print(f"Preços distribuição: {count} linhas")
    if count < PRECOS_DIST_MIN_ROWS:
        raise RuntimeError(
            f"Preços de distribuição com {count} linhas (mínimo {PRECOS_DIST_MIN_ROWS}) — abortando"
        )
    return ("\n".join(lines) + "\n").encode("utf-8")


def fetch_precos_produtores() -> bytes:
    """Preços semanais de PRODUTORES/importadores por região (desde 2013):
    INICIO;FIM;REGIAO;PRODUTO;UNIDADE;PRECO. Publica apenas os produtos de
    interesse do consumidor (PRODUTORES_PRODUTOS); valores sigilosos vêm
    como "***" na planilha e são descartados."""
    print("Preços produtores: baixando...")
    book = xlrd.open_workbook(file_contents=download(PRECOS_PRODUTORES_URL))
    sheet = book.sheet_by_index(0)

    # Colunas fixas da planilha: produto, início, fim, 5 regiões e Brasil.
    regioes = ["NORTE", "NORDESTE", "CENTRO-OESTE", "SUL", "SUDESTE", "BRASIL"]
    lines = ["INICIO;FIM;REGIAO;PRODUTO;UNIDADE;PRECO"]
    count = 0

    for r in range(sheet.nrows):
        row = sheet.row_values(r)
        if len(row) < 9:
            continue
        produto_raw = str(row[0]).strip()
        if not produto_raw.startswith(PRODUTORES_PRODUTOS):
            continue
        try:
            inicio = xlrd.xldate_as_datetime(row[1], book.datemode)
            fim = xlrd.xldate_as_datetime(row[2], book.datemode)
        except (TypeError, ValueError, xlrd.xldate.XLDateError):
            continue

        # "Gasolina A Comum (R$/litro)" → nome limpo + unidade separada.
        nome, unidade = produto_raw, ""
        if "(" in produto_raw and produto_raw.endswith(")"):
            nome, _, resto = produto_raw.partition("(")
            nome = nome.strip().upper()
            unidade = resto[:-1].strip()

        for regiao, valor in zip(regioes, row[3:9]):
            if isinstance(valor, (int, float)) and valor > 0:
                lines.append(";".join([
                    inicio.strftime("%Y-%m-%d"),
                    fim.strftime("%Y-%m-%d"),
                    regiao,
                    nome,
                    unidade,
                    f"{float(valor):.4f}",
                ]))
                count += 1

    print(f"Preços produtores: {count} linhas")
    if count < PRECOS_PROD_MIN_ROWS:
        raise RuntimeError(
            f"Preços de produtores com {count} linhas (mínimo {PRECOS_PROD_MIN_ROWS}) — abortando"
        )
    return ("\n".join(lines) + "\n").encode("utf-8")


def main() -> None:
    print(f"Baixando {SOURCE_URL}")
    payload = download(SOURCE_URL)
    print(f"Download concluído: {len(payload)/1e6:.1f} MB")
    if len(payload) < 1_000_000:
        raise RuntimeError(f"Arquivo suspeito de {len(payload)} bytes — esperado > 1 MB")

    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    sheet = workbook.worksheets[0]

    rows_iter = sheet.iter_rows(values_only=True)
    header = [str(cell).strip() if cell is not None else "" for cell in next(rows_iter)]
    if header != EXPECTED_HEADER:
        raise RuntimeError(
            "Header do xlsx mudou — atualizar o app antes de publicar.\n"
            f"Esperado: {EXPECTED_HEADER}\nRecebido: {header}"
        )

    date_index = EXPECTED_HEADER.index("DATA DO DF")
    doc_index = EXPECTED_HEADER.index("CNPJ/CPF")
    buffer = io.StringIO()
    writer = csv.writer(buffer, quoting=csv.QUOTE_MINIMAL, lineterminator="\n")
    writer.writerow(EXPECTED_HEADER)

    count = 0
    for row in rows_iter:
        if row is None or all(cell is None for cell in row):
            continue
        cells = list(row[: len(EXPECTED_HEADER)])
        cells += [None] * (len(EXPECTED_HEADER) - len(cells))
        formatted = [
            format_cell(cell, index == date_index) for index, cell in enumerate(cells)
        ]
        formatted[doc_index] = mask_cpf(formatted[doc_index])
        writer.writerow(formatted)
        count += 1

    print(f"Convertidas {count} linhas")
    if count < MIN_ROWS:
        raise RuntimeError(f"Apenas {count} linhas (mínimo esperado {MIN_ROWS}) — abortando")

    csv_bytes = buffer.getvalue().encode("utf-8")
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    (DATA_DIR / "base_anp.csv").write_bytes(csv_bytes)

    cadastro_bytes = fetch_cadastro()
    (DATA_DIR / "cadastro_postos.csv").write_bytes(cadastro_bytes)

    # Procon-SP falha SUAVE: se o portal estiver fora do ar, mantém a última
    # versão publicada em vez de derrubar a atualização da ANP inteira.
    cadastro_cnpjs = {
        line.split(";", 1)[0]
        for line in cadastro_bytes.decode("utf-8").splitlines()[1:]
    }
    try:
        procon_bytes = fetch_procon_sp(cadastro_cnpjs)
    except Exception as error:  # noqa: BLE001
        print(f"AVISO: Procon-SP indisponível ({error}); reutilizando a última versão")
        try:
            procon_bytes = download(
                "https://github.com/LuisFontinelles/octano-anp-data/releases/download/"
                "latest/procon_sp.csv",
                attempts=2,
            )
        except Exception as fallback_error:  # noqa: BLE001
            # Nem coleta nem fallback: publica só o header e OMITE o sha do
            # metadata — o app mantém o que já tem e nada derruba a ANP.
            print(f"AVISO: fallback do Procon-SP também falhou ({fallback_error})")
            procon_bytes = None
    (DATA_DIR / "procon_sp.csv").write_bytes(
        procon_bytes or b"CNPJ;RAZAO_SOCIAL;PROCESSO;STATUS;VALOR_MULTA\n"
    )

    # Reclamações Fundamentadas (PowerBI): falha SUAVE, mesmo padrão do Procon.
    try:
        reclamadas_bytes = fetch_reclamadas_procon_sp(cadastro_cnpjs)
    except Exception as error:  # noqa: BLE001
        print(f"AVISO: Reclamadas Procon-SP indisponível ({error}); reutilizando a última versão")
        try:
            reclamadas_bytes = download(
                "https://github.com/LuisFontinelles/octano-anp-data/releases/download/"
                "latest/reclamadas_procon_sp.csv",
                attempts=2,
            )
        except Exception as fallback_error:  # noqa: BLE001
            print(f"AVISO: fallback das Reclamadas também falhou ({fallback_error})")
            reclamadas_bytes = None
    (DATA_DIR / "reclamadas_procon_sp.csv").write_bytes(
        reclamadas_bytes or b"CNPJ\n"
    )

    # IPEM Bombas Seguras (PowerBI): falha SUAVE, mesmo padrão do Procon.
    try:
        ipem_bytes = fetch_ipem_bombas_seguras()
    except Exception as error:  # noqa: BLE001
        print(f"AVISO: IPEM Bombas Seguras indisponível ({error}); reutilizando a última versão")
        try:
            ipem_bytes = download(
                "https://github.com/LuisFontinelles/octano-anp-data/releases/download/"
                "latest/ipem_bombas_seguras.json",
                attempts=2,
            )
        except Exception as fallback_error:  # noqa: BLE001
            print(f"AVISO: fallback do IPEM também falhou ({fallback_error})")
            ipem_bytes = None
    (DATA_DIR / "ipem_bombas_seguras.json").write_bytes(ipem_bytes or b"[]\n")

    # Reclamações Fundamentadas SINDEC (nacional): falha SUAVE, mesmo padrão.
    try:
        sindec_bytes = fetch_reclamadas_sindec()
    except Exception as error:  # noqa: BLE001
        print(f"AVISO: SINDEC indisponível ({error}); reutilizando a última versão")
        try:
            sindec_bytes = download(
                "https://github.com/LuisFontinelles/octano-anp-data/releases/download/"
                "latest/reclamadas_sindec.csv",
                attempts=2,
            )
        except Exception as fallback_error:  # noqa: BLE001
            print(f"AVISO: fallback do SINDEC também falhou ({fallback_error})")
            sindec_bytes = None
    (DATA_DIR / "reclamadas_sindec.csv").write_bytes(
        sindec_bytes or b"CNPJ;RAZAO_SOCIAL;UFS;ATENDIDAS;NAO_ATENDIDAS;TOTAL;RESOLUTIVIDADE\n"
    )

    precos_dist_bytes = fetch_precos_distribuicao()
    (DATA_DIR / "precos_distribuicao.csv").write_bytes(precos_dist_bytes)

    precos_prod_bytes = fetch_precos_produtores()
    (DATA_DIR / "precos_produtores.csv").write_bytes(precos_prod_bytes)

    now = datetime.now(timezone.utc)
    metadata = {
        "updatedAt": now.strftime("%Y-%m-%d"),
        "updatedAtISO": now.isoformat(timespec="seconds"),
        "rows": count,
        "sha256": hashlib.sha256(csv_bytes).hexdigest(),
        "source": SOURCE_URL,
        "cadastroRows": cadastro_bytes.count(b"\n") - 1,
        "cadastroSha256": hashlib.sha256(cadastro_bytes).hexdigest(),
        "cadastroSource": "https://revendedoresapi.anp.gov.br/v1/combustivel",
        "precosDistRows": precos_dist_bytes.count(b"\n") - 1,
        "precosDistSha256": hashlib.sha256(precos_dist_bytes).hexdigest(),
        "precosProdRows": precos_prod_bytes.count(b"\n") - 1,
        "precosProdSha256": hashlib.sha256(precos_prod_bytes).hexdigest(),
        "precosSource": "https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/precos",
    }
    if procon_bytes:
        metadata["proconRows"] = procon_bytes.count(b"\n") - 1
        metadata["proconSha256"] = hashlib.sha256(procon_bytes).hexdigest()
        metadata["proconSource"] = (
            "https://sistemas.procon.sp.gov.br/transparencia/empresas_autuadas/"
        )
    if reclamadas_bytes:
        metadata["reclamadasRows"] = reclamadas_bytes.count(b"\n") - 1
        metadata["reclamadasSha256"] = hashlib.sha256(reclamadas_bytes).hexdigest()
        metadata["reclamadasSource"] = (
            "https://www.procon.sp.gov.br/empresas-reclamadas/"
        )
    if ipem_bytes:
        # JSON array de CNPJs: a contagem é o tamanho do array, não linhas.
        metadata["ipemRows"] = len(json.loads(ipem_bytes))
        metadata["ipemSha256"] = hashlib.sha256(ipem_bytes).hexdigest()
        metadata["ipemSource"] = "https://www.ipem.sp.gov.br/bombasegura"
    if sindec_bytes:
        metadata["sindecRows"] = sindec_bytes.count(b"\n") - 1
        metadata["sindecSha256"] = hashlib.sha256(sindec_bytes).hexdigest()
        metadata["sindecSource"] = (
            "https://dados.mj.gov.br/dataset/"
            "cadastro-nacional-de-reclamacoes-fundamentadas-procons-sindec"
        )
    (DATA_DIR / "metadata.json").write_text(
        json.dumps(metadata, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    print(
        f"OK: base_anp.csv ({len(csv_bytes)/1e6:.1f} MB), "
        f"cadastro_postos.csv ({len(cadastro_bytes)/1e6:.1f} MB), "
        f"precos_distribuicao.csv ({len(precos_dist_bytes)/1e6:.1f} MB), "
        f"precos_produtores.csv ({len(precos_prod_bytes)/1e6:.1f} MB), "
        f"procon_sp.csv ({len(procon_bytes or b'')/1e3:.0f} KB), "
        f"reclamadas_procon_sp.csv ({len(reclamadas_bytes or b'')/1e3:.0f} KB), "
        f"ipem_bombas_seguras.json ({len(ipem_bytes or b'')/1e3:.0f} KB), "
        f"reclamadas_sindec.csv ({len(sindec_bytes or b'')/1e3:.0f} KB) "
        f"e metadata.json gravados"
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as error:  # noqa: BLE001 - queremos falhar o workflow com mensagem clara
        print(f"ERRO: {error}", file=sys.stderr)
        sys.exit(1)
